"""QCOS 2026年8月历史组局/收银数据迁移导入器

从 Excel「原始组局记录」导入 2026-08-01 00:00:00 ~ 2026-08-20 23:59:59 的
历史 sessions / session_players / payments / visit_records，复用现有 QCOS
数据结构，不做第二套业务表。

设计原则：
- 幂等：sessions.source_id = excel_<date>_<table_no>（与既有约定一致）+ 
  legacy_import_records 登记表 UNIQUE(source, source_sheet, source_row)，同一
  Excel 重跑任意次，库里仍只有一份历史数据。
- 不猜：支付金额只取 Excel 有效数字；支付方式统一 legacy_unknown（不猜现金/
  微信/支付宝）；时间只精确到日期（time_precision='date_only'），不伪造具体
  开台/结账时刻。
- 不合并：玩家匹配只做「标准昵称 trim 后 exact unique（含 ASCII 大小写不敏感
  唯一命中，BACK==back）」或「Excel 唯一玩家序号命中 players.qcos_id」两种精确
  匹配；重名（>1）列入 AMBIGUOUS，绝不自动绑定、绝不相似名合并。
- 安全：--execute 前必须备份成功；全部写入在单事务内，任一关键校验失败回滚。

用法：
  python scripts/import_august_2026_history.py \\
      --excel "QCOS_青鹭玩家数据库_V2.0 - 副本.xlsx" --db qcos.db --dry-run
  python scripts/import_august_2026_history.py \\
      --excel "QCOS_青鹭玩家数据库_V2.0 - 副本.xlsx" --db qcos.db --execute

可选：
  --report-dir reports/aug2026_import
  --create-missing-players    （默认不创建缺失玩家；启用后按 exact 标准昵称创建，
                                空昵称 / #REF! 永不创建）
  --allow-legacy-gap          （放行 8/1-8/13 旧导入与 Excel 的 GMV 缺口，默认回滚）
  --ambiguous-bind 昵称:player_id,...  （人工确认的重名绑定，如 back:84,坤哥:85）
"""

import argparse
import csv
import hashlib
import json
import os
import sys
import sqlite3
import uuid
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from backup_qcos import backup_database  # noqa: E402

SOURCE = 'legacy_excel_aug2026'
SHEET_NAME = '原始组局记录'
DATE_START = date(2026, 8, 1)
DATE_END = date(2026, 8, 21)  # 含头不含尾：>= 8/1 且 < 8/21

# Excel 表头 -> 内部字段名
COLUMN_ALIASES = {
    '日期': 'date',
    '姓名': 'raw_name',
    '四/八': 'machine_type',
    '娱乐/竞技': 'game_type',
    '是否带人': 'brought_guest',
    '组织者': 'organizer',
    '是否通宵': 'is_overnight',
    '标准昵称': 'nickname',
    '唯一玩家序号': 'player_seq',
    '牌桌序号': 'table_no',
    '桌首标记': 'is_table_head',
    '桌首组织者': 'table_head_organizer',
    '数据质量': 'data_quality',
    '真实支付金额': 'payment_amount',
}

# 机器分配：8口机 [1,2]，4口机 [3,4]（与 import_sessions_from_excel.py 一致）
EIGHT_PORT_IDS = [1, 2]
FOUR_PORT_IDS = [3, 4]


# ---------------------------------------------------------------------------
# 解析辅助
# ---------------------------------------------------------------------------

def parse_machine(value):
    """Excel 四/八 -> '8port' / '4port' / None"""
    if value is None:
        return None
    s = str(value).strip().lower()
    if s in ('8', '八', '八口', '8口', '八麻', '8port'):
        return '8port'
    if s in ('4', '四', '四口', '4口', '四麻', '4port'):
        return '4port'
    return None


def parse_overnight(value):
    """是否通宵：'是'/'1'/true -> 1，'否'/'' -> 0"""
    if value is None:
        return 0
    s = str(value).strip().lower()
    return 1 if s in ('是', '1', 'true', 'yes', 'y') else 0


def parse_yes_no(value, default=0):
    if value is None:
        return default
    s = str(value).strip().lower()
    if s in ('是', '1', 'true', 'yes', 'y'):
        return 1
    if s in ('否', '0', 'false', 'no', 'n'):
        return 0
    return default


def parse_amount(value):
    """真实支付金额 -> (amount: float|None, remark: str|None)
    - 空 -> (None, None)  不猜
    - 数字/数字字符串 -> (float, None)
    - 其他文字（免单券等）-> (None, 原文)
    """
    if value is None:
        return None, None
    s = str(value).strip()
    if not s:
        return None, None
    try:
        return float(s), None
    except ValueError:
        return None, s


def parse_table_no(value):
    """牌桌序号 -> (int|None, err: str|None)"""
    if value is None:
        return None, '空牌桌序号'
    s = str(value).strip()
    if s in ('', '#REF!', '#REF'):
        return None, f'牌桌序号非法:{s}'
    try:
        return int(float(s)), None
    except (ValueError, TypeError):
        return None, f'牌桌序号非法:{s}'


def parse_date_cell(value):
    """日期列 -> date|None"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d'):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def row_hash(row):
    """对行内容做 sha1，用于 legacy_import_records.hash"""
    payload = json.dumps(row, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha1(payload.encode('utf-8')).hexdigest()[:16]


# ---------------------------------------------------------------------------
# Excel 读取
# ---------------------------------------------------------------------------

class ExcelLoader:
    def __init__(self, path):
        self.path = path
        try:
            import openpyxl
        except ImportError as ex:
            raise SystemExit('缺少 openpyxl，请先安装: pip install openpyxl') from ex
        self.wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if SHEET_NAME not in self.wb.sheetnames:
            self.wb.close()
            raise SystemExit(f'Excel 缺少工作表「{SHEET_NAME}」，实际: {self.wb.sheetnames}')
        self.ws = self.wb[SHEET_NAME]
        self.headers = None
        self.col_index = {}
        self._read_headers()

    def _read_headers(self):
        for row in self.ws.iter_rows(min_row=1, max_row=1, values_only=True):
            self.headers = row
            break
        if not self.headers:
            raise SystemExit('Excel 首行为空，无法识别表头')
        for i, h in enumerate(self.headers):
            key = str(h).strip() if h is not None else ''
            if key in COLUMN_ALIASES and key not in self.col_index:
                self.col_index[COLUMN_ALIASES[key]] = i

    def iter_rows(self):
        """yield (source_row, dict)。source_row 为 Excel 行号（1-based，含表头行）。
        只产出 8/1-8/20 范围内的行。"""
        for n, row in enumerate(self.ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(v is None for v in row):
                continue
            rec = {field: (row[i] if i < len(row) else None)
                   for field, i in self.col_index.items()}
            d = parse_date_cell(rec.get('date'))
            if d is None:
                continue
            if not (DATE_START <= d < DATE_END):
                continue
            rec['date'] = d  # 统一为 date 对象，避免 datetime 混入 source_id
            rec['_source_row'] = n
            yield n, rec

    def close(self):
        try:
            self.wb.close()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# 玩家匹配
# ---------------------------------------------------------------------------

class PlayerMatcher:
    """精确匹配，禁止模糊/相似名合并。

    优先级：
      1. Excel 唯一玩家序号 N -> players.qcos_id 数字部分 == N（唯一命中）
      2. 标准昵称 trim 后 exact unique match
      3. 0 个 -> unmatched；多于 1 个 -> ambiguous（不猜）
    """

    def __init__(self, db, whitelist=None):
        self.qcos_seq_to_pid = {}
        self.by_name = defaultdict(list)
        self.by_name_lower = defaultdict(list)  # ASCII 大小写不敏感兜底（BACK==back）
        # 人工确认白名单：昵称(大小写不敏感) -> player_id。业务侧已确认的唯一绑定，
        # 用于重名(ambiguous)玩家的人工拍板；仍属精确绑定，不是自动猜测。
        self.whitelist_lower = {
            str(k).strip().lower(): int(v) for k, v in (whitelist or {}).items()
        }
        # 兼容未迁移库（无 status 列时不过滤；已迁移库排除 archived）
        cols = [r['name'] for r in db.execute('PRAGMA table_info(players)')]
        if 'status' in cols:
            rows = db.execute(
                "SELECT id, name, qcos_id FROM players WHERE status IS NULL OR status='active'"
            ).fetchall()
        else:
            rows = db.execute("SELECT id, name, qcos_id FROM players").fetchall()
        for r in rows:
            pid, name, qcos_id = r['id'], (r['name'] or '').strip(), r['qcos_id']
            if qcos_id:
                s = str(qcos_id).strip()
                if s.startswith('P') and s[1:].isdigit():
                    seq = int(s[1:])
                    self.qcos_seq_to_pid.setdefault(seq, pid)
            if name:
                self.by_name[name].append(pid)
                self.by_name_lower[name.lower()].append(pid)

    def match(self, nickname, player_seq):
        """返回 (player_id|None, status) status in exact/qcos_seq/whitelist/unmatched/ambiguous

        精确匹配语义：先按 Excel 唯一玩家序号命中 qcos_id；再查人工确认白名单；
        再按原样唯一命中；未命中时按 ASCII 大小写不敏感唯一命中（BACK==back，
        仍是唯一精确匹配，不引入模糊/相似名合并）。
        """
        nick = (nickname or '').strip()
        if player_seq is not None:
            pid = self.qcos_seq_to_pid.get(int(player_seq))
            if pid is not None:
                return pid, 'qcos_seq'
        if not nick:
            return None, 'unmatched'
        wl = self.whitelist_lower.get(nick.lower())
        if wl is not None:
            return wl, 'whitelist'
        cands = self.by_name.get(nick, [])
        if len(cands) == 1:
            return cands[0], 'exact'
        if len(cands) > 1:
            return None, 'ambiguous'
        # 原样未命中 -> 大小写不敏感兜底（中文 lower() 无变化，不受影响）
        cands2 = self.by_name_lower.get(nick.lower(), [])
        if len(cands2) == 1:
            return cands2[0], 'exact'
        if len(cands2) > 1:
            return None, 'ambiguous'
        return None, 'unmatched'


# ---------------------------------------------------------------------------
# 导入器
# ---------------------------------------------------------------------------

class Importer:
    def __init__(self, excel_path, db_path, report_dir, create_missing=False,
                 allow_legacy_gap=False, backup_dir=None, ambiguous_bind=None):
        self.excel_path = excel_path
        self.db_path = db_path
        self.report_dir = report_dir
        self.create_missing = create_missing
        self.allow_legacy_gap = allow_legacy_gap
        self.backup_dir = backup_dir or os.path.join(ROOT, 'backups')
        self.ambiguous_bind = ambiguous_bind or {}
        self.loader = ExcelLoader(excel_path)

        # 只读连接用于匹配/分析
        self.db = sqlite3.connect(f'file:{db_path}?mode=ro', uri=True)
        self.db.row_factory = sqlite3.Row
        self.matcher = PlayerMatcher(self.db, whitelist=self.ambiguous_bind)

    # ---- 清洗 ----
    def clean_row(self, n, rec):
        """返回 (clean_rec, hard_issues, soft_issues)。

        hard_issues：行无法归属（昵称为空 / 牌桌非法）-> 整行跳过
        soft_issues：行保留，但标记（金额缺失、数据质量=#REF!、机型异常等）
        """
        hard_issues, soft_issues = [], []
        nick = str(rec.get('nickname') or '').strip()
        if not nick:
            hard_issues.append('标准昵称为空')
        elif nick == '#REF!':
            hard_issues.append('标准昵称为#REF!')

        tbl, tbl_err = parse_table_no(rec.get('table_no'))
        if tbl_err:
            hard_issues.append(tbl_err)

        mt = parse_machine(rec.get('machine_type'))
        if mt is None:
            soft_issues.append(f'机型非法:{rec.get("machine_type")!r}')

        dq = str(rec.get('data_quality') or '').strip()
        if dq == '#REF!':
            soft_issues.append('数据质量=#REF!')

        amount, remark = parse_amount(rec.get('payment_amount'))
        if amount is None:
            soft_issues.append(f'支付金额缺失/非数字:{rec.get("payment_amount")!r}')

        seq = rec.get('player_seq')
        seq_int = None
        if seq is not None and str(seq).strip() not in ('', '#REF!'):
            try:
                seq_int = int(float(seq))
            except (ValueError, TypeError):
                soft_issues.append(f'唯一玩家序号非法:{seq!r}')

        clean = {
            'source_row': n,
            'date': rec['date'],
            'raw_name': str(rec.get('raw_name') or '').strip(),
            'nickname': nick,
            'machine_type': mt,
            'game_type': str(rec.get('game_type') or '').strip(),
            'brought_guest': parse_yes_no(rec.get('brought_guest')),
            'organizer': str(rec.get('organizer') or '').strip(),
            'is_overnight': parse_overnight(rec.get('is_overnight')),
            'player_seq': seq_int,
            'table_no': tbl,
            'is_table_head': parse_yes_no(rec.get('is_table_head')),
            'table_head_organizer': str(rec.get('table_head_organizer') or '').strip(),
            'data_quality': dq,
            'amount': amount,
            'amount_remark': remark,
        }
        return clean, hard_issues, soft_issues

    def analyze(self):
        """全量解析与匹配，不写库。返回分析结果 dict。"""
        stats = {
            'excel_rows': 0,
            'valid_player_rows': 0,
            'bad_rows': [],
            'missing_payments': [],
            'unmatched_players': {},
            'ambiguous_players': {},
            'payments_detected': 0,
            'excel_gmv': 0.0,
            'partial_sessions': 0,
        }
        tables = defaultdict(list)   # (date, table_no) -> [clean_row]
        bad_table_keys = []          # 牌桌非法无法归组的行

        for n, rec in self.loader.iter_rows():
            stats['excel_rows'] += 1
            clean, hard_issues, soft_issues = self.clean_row(n, rec)
            if hard_issues:
                stats['bad_rows'].append({
                    'source_row': n, 'date': str(rec['date']),
                    'nickname': str(rec.get('nickname') or '').strip(),
                    'issues': hard_issues + soft_issues,
                })
                continue

            # 归属桌
            if clean['table_no'] is None:
                bad_table_keys.append(clean)
                continue

            stats['valid_player_rows'] += 1
            # 玩家匹配
            pid, mstatus = self.matcher.match(clean['nickname'], clean['player_seq'])
            clean['player_id'] = pid
            clean['match_status'] = mstatus
            if mstatus == 'unmatched':
                stats['unmatched_players'].setdefault(clean['nickname'], []).append(clean['source_row'])
            elif mstatus == 'ambiguous':
                stats['ambiguous_players'].setdefault(clean['nickname'], []).append(clean['source_row'])
            elif mstatus == 'qcos_seq':
                stats.setdefault('matched_qcos_seq', 0)
                stats['matched_qcos_seq'] += 1

            if clean['amount'] is None:
                stats['missing_payments'].append({
                    'source_row': n, 'date': str(clean['date']),
                    'nickname': clean['nickname'], 'table_no': clean['table_no'],
                    'remark': clean['amount_remark'],
                })
            else:
                stats['payments_detected'] += 1
                stats['excel_gmv'] = round(stats['excel_gmv'] + clean['amount'], 2)

            tables[(clean['date'], clean['table_no'])].append(clean)

        # 桌分组统计
        session_groups = []
        for key, rows in sorted(tables.items(), key=lambda kv: (str(kv[0][0]), kv[0][1])):
            known = [r for r in rows if r['nickname']]
            is_partial = len(known) < 4
            if is_partial:
                stats['partial_sessions'] += 1
            # 机型取非空众数
            mt_counter = Counter(r['machine_type'] for r in rows if r['machine_type'])
            machine_type = mt_counter.most_common(1)[0][0] if mt_counter else None
            session_groups.append({
                'date': key[0], 'table_no': key[1],
                'rows': rows, 'is_partial': is_partial,
                'machine_type': machine_type,
                'is_overnight': 1 if any(r['is_overnight'] for r in rows) else 0,
            })

        stats['sessions_detected'] = len(session_groups)
        stats['session_groups'] = session_groups
        stats['bad_table_keys'] = bad_table_keys

        # 幂等检测：已导入 session（source_id 命中）
        existing = set()
        for r in self.db.execute("SELECT source_id FROM sessions WHERE source_id IS NOT NULL"):
            existing.add(r['source_id'])
        stats['existing_aug_sessions'] = sum(
            1 for sid in existing if sid.startswith('excel_2026-08-')
        )
        already = 0
        would = []
        no_machine = 0
        for g in session_groups:
            sid = f"excel_{g['date']}_{g['table_no']}"
            g['source_id'] = sid
            if g['machine_type'] is None:
                # 整桌机型无法确定：不猜，跳过并报告
                g['status'] = 'SKIP_NO_MACHINE_TYPE'
                no_machine += 1
            elif sid in existing:
                already += 1
                g['status'] = 'SKIP_ALREADY_IMPORTED'
            else:
                g['status'] = 'WOULD_INSERT'
                would.append(g)
        stats['already_imported'] = already
        stats['would_insert_sessions'] = len(would)
        stats['skipped_no_machine'] = no_machine
        stats['would_insert_groups'] = would

        # 行级幂等：legacy_import_records 命中（未迁移库无此表时视为空）
        imported_rows = set()
        has_reg = self.db.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='legacy_import_records'"
        ).fetchone()
        if has_reg:
            for r in self.db.execute(
                "SELECT source_row FROM legacy_import_records WHERE source=?", [SOURCE]
            ):
                imported_rows.add(r['source_row'])
        stats['row_level_already_imported'] = len(imported_rows)

        # 估算本次新增的 GMV（would_insert 桌的有效金额）
        stats['would_insert_gmv'] = round(
            sum(r['amount'] or 0 for g in would for r in g['rows'] if r['amount']), 2
        )
        stats['would_insert_payments'] = sum(
            1 for g in would for r in g['rows'] if r['amount'] is not None
        )
        stats['would_insert_session_players'] = sum(len(g['rows']) for g in would)

        # 已导入部分（8/1-8/20 已存在 source_id 桌）的库内 GMV 与 Excel GMV
        stats['existing_imported_gmv'] = self._existing_gmv_for_aug()
        stats['projected_gmv'] = round(stats['existing_imported_gmv'] + stats['would_insert_gmv'], 2)
        stats['gmv_diff'] = round(stats['excel_gmv'] - stats['projected_gmv'], 2)

        return stats

    def _existing_gmv_for_aug(self):
        total = 0.0
        rows = self.db.execute(
            """SELECT p.amount FROM payments p
               JOIN session_players sp ON p.session_player_id = sp.id
               JOIN sessions s ON sp.session_id = s.id
               WHERE s.source_id LIKE 'excel_2026-08-%' AND p.amount IS NOT NULL"""
        ).fetchall()
        return round(sum(r['amount'] for r in rows), 2)

    # ---- 报告 ----
    def write_reports(self, stats):
        os.makedirs(self.report_dir, exist_ok=True)
        p = lambda name: os.path.join(self.report_dir, name)  # noqa: E731

        summary = {
            'source': SOURCE,
            'excel_file': os.path.basename(self.excel_path),
            'date_start': str(DATE_START),
            'date_end': str(DATE_END - timedelta(days=1)),
            'excel_rows': stats['excel_rows'],
            'valid_player_rows': stats['valid_player_rows'],
            'sessions_detected': stats['sessions_detected'],
            'unique_players': len(set(r['nickname'] for g in stats['session_groups'] for r in g['rows'])),
            'matched_players': len(set(r['nickname'] for g in stats['session_groups'] for r in g['rows']
                                        if r['match_status'] in ('exact', 'qcos_seq'))),
            'unmatched_players': len(stats['unmatched_players']),
            'ambiguous_players': len(stats['ambiguous_players']),
            'payments_detected': stats['payments_detected'],
            'excel_gmv': stats['excel_gmv'],
            'partial_sessions': stats['partial_sessions'],
            'bad_rows': len(stats['bad_rows']),
            'would_insert_sessions': stats['would_insert_sessions'],
            'would_insert_session_players': stats['would_insert_session_players'],
            'would_insert_payments': stats['would_insert_payments'],
            'already_imported': stats['already_imported'],
            'skipped_no_machine': stats['skipped_no_machine'],
            'existing_aug_sessions': stats['existing_aug_sessions'],
            'existing_imported_gmv': stats['existing_imported_gmv'],
            'would_insert_gmv': stats['would_insert_gmv'],
            'projected_gmv': stats['projected_gmv'],
            'gmv_diff': stats['gmv_diff'],
        }
        with open(p('summary.json'), 'w', encoding='utf-8') as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)

        with open(p('tables.csv'), 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.writer(f)
            w.writerow(['日期', '牌桌序号', 'source_id', '机型', '通宵', '状态', '玩家数',
                        '有支付人数', '桌金额'])
            for g in stats['session_groups']:
                amt = sum(r['amount'] or 0 for r in g['rows'])
                w.writerow([g['date'], g['table_no'], g['source_id'], g['machine_type'],
                            g['is_overnight'], g['status'], len(g['rows']),
                            sum(1 for r in g['rows'] if r['amount'] is not None),
                            round(amt, 2)])

        with open(p('players.csv'), 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.writer(f)
            w.writerow(['标准昵称', '唯一玩家序号', '匹配player_id', '匹配方式', '行数', '金额合计'])
            aggr = defaultdict(lambda: {'rows': 0, 'amt': 0.0, 'seq': None, 'pid': None, 'ms': None})
            for g in stats['session_groups']:
                for r in g['rows']:
                    a = aggr[r['nickname']]
                    a['rows'] += 1
                    a['amt'] += r['amount'] or 0
                    a['seq'] = r['player_seq']
                    a['pid'] = r['player_id']
                    a['ms'] = r['match_status']
            for nick, a in sorted(aggr.items()):
                w.writerow([nick, a['seq'], a['pid'], a['ms'], a['rows'], round(a['amt'], 2)])

        with open(p('unmatched_players.csv'), 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.writer(f)
            w.writerow(['标准昵称', 'Excel行号'])
            for nick, rows in sorted(stats['unmatched_players'].items()):
                for n in rows:
                    w.writerow([nick, n])

        with open(p('ambiguous_players.csv'), 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.writer(f)
            w.writerow(['标准昵称', 'Excel行号', 'QCOS候选玩家(id:name)'])
            cands = {}
            for r in self.db.execute(
                "SELECT id, name, qcos_id FROM players WHERE name IN (SELECT DISTINCT name FROM players)"
            ):
                pass
            # 直接按昵称查候选
            for nick in stats['ambiguous_players']:
                cs = self.db.execute(
                    'SELECT id, name, qcos_id FROM players WHERE name=?', [nick]
                ).fetchall()
                cands[nick] = [f"{c['id']}:{c['name']}({c['qcos_id'] or '无'})" for c in cs]
            for nick, rows in sorted(stats['ambiguous_players'].items()):
                for n in rows:
                    w.writerow([nick, n, '; '.join(cands.get(nick, []))])

        with open(p('missing_payments.csv'), 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.writer(f)
            w.writerow(['Excel行号', '日期', '标准昵称', '牌桌序号', '原始金额', '说明'])
            for m in stats['missing_payments']:
                w.writerow([m['source_row'], m['date'], m['nickname'], m['table_no'],
                            m['remark'] or '(空)', 'PAYMENT_MISSING'])

        with open(p('bad_rows.csv'), 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.writer(f)
            w.writerow(['Excel行号', '日期', '标准昵称', '问题'])
            for b in stats['bad_rows']:
                w.writerow([b['source_row'], b['date'], b['nickname'], '; '.join(b['issues'])])

        return summary

    # ---- 正式导入 ----
    def _ensure_schema(self, conn):
        """幂等补齐历史导入所需 schema（与 models.init_db 迁移等价，但作用于 --db 目标库）。

        未迁移库（真实 qcos.db / 副本）自动补 legacy_import_records 表 + sessions
        两列；已迁移库无操作。须在备份成功后调用。
        """
        conn.execute(
            '''CREATE TABLE IF NOT EXISTS legacy_import_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source TEXT NOT NULL,
                source_sheet TEXT,
                source_row INTEGER,
                legacy_table_no TEXT,
                date TEXT,
                entity_type TEXT,
                entity_id INTEGER,
                imported_at TEXT,
                hash TEXT,
                UNIQUE(source, source_sheet, source_row)
            )'''
        )
        s_cols = [r[1] for r in conn.execute('PRAGMA table_info(sessions)')]
        if 'time_precision' not in s_cols:
            conn.execute('ALTER TABLE sessions ADD COLUMN time_precision TEXT')
        if 'import_quality' not in s_cols:
            conn.execute('ALTER TABLE sessions ADD COLUMN import_quality TEXT')
        conn.commit()

    def execute(self):
        """备份 -> schema 迁移 -> 事务内写入 -> 校验 -> COMMIT/ROLLBACK。返回 (stats, inserted, ok, checks)"""
        # 0) 备份检查
        try:
            backup_database(db_path=self.db_path, backup_dir=self.backup_dir)
        except Exception as ex:
            return None, None, False, f'备份失败，禁止导入: {ex}'
        print(f'[backup] 备份完成 -> {self.backup_dir}')

        # 0.5) schema 迁移（幂等；未迁移库先补齐 legacy_import_records / sessions 列）
        with sqlite3.connect(self.db_path) as conn:
            self._ensure_schema(conn)
        print('[schema] legacy_import_records / sessions.time_precision / import_quality 就绪')

        stats = self.analyze()
        # 基线：8/21 以后的真实生产数据（导入绝不允许触碰）
        stats['_baseline_post_821'] = self.db.execute(
            """SELECT COUNT(*) c FROM sessions
               WHERE start_time >= '2026-08-21' AND source_id IS NULL"""
        ).fetchone()['c']
        if stats['gmv_diff'] != 0:
            # 已有数据与 Excel 不一致（8/1-8/13 历史导入缺口）——不允许覆盖，提示人工
            print(f'[warn] 已导入部分与 Excel GMV 差异 {stats["gmv_diff"]} 元，'
                  f'本次仅能保证新增部分正确。')

        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        try:
            conn.execute('BEGIN')
            now = datetime.now().isoformat(timespec='seconds')
            inserted = {'sessions': 0, 'session_players': 0, 'payments': 0,
                        'visit_records': 0, 'import_records': 0, 'skipped_rows': 0}

            # --create-missing-players：事务内为 unmatched 昵称建档并绑定（仅合法昵称）
            created_players = {}
            if self.create_missing:
                for nick in stats['unmatched_players']:
                    if not nick or '#REF!' in nick:
                        continue
                    cur = conn.execute(
                        'INSERT INTO players (name, created_at, updated_at) VALUES (?, ?, ?)',
                        [nick, now, now]
                    )
                    created_players[nick] = cur.lastrowid
                    print(f'  [create_player] {nick} -> id={cur.lastrowid}')
                if created_players:
                    for g in stats['session_groups']:
                        for r in g['rows']:
                            if r['player_id'] is None and r['nickname'] in created_players:
                                r['player_id'] = created_players[r['nickname']]
                                r['match_status'] = 'created'
                    stats['created_players'] = created_players

            # 已导入行（防止同 Excel 行重复处理）
            imported_rows = set(r[0] for r in conn.execute(
                'SELECT source_row FROM legacy_import_records WHERE source=?', [SOURCE]
            ))
            day_counter = defaultdict(int)

            for g in stats['session_groups']:
                if g['status'] == 'SKIP_NO_MACHINE_TYPE':
                    # 整桌机型无法确定：不猜，不导入，不登记（修正 Excel 后可重跑）
                    continue
                if g['status'] == 'SKIP_ALREADY_IMPORTED':
                    # 桌级已导入：行级登记确保不重复（若行尚未登记则登记为 skip）
                    for r in g['rows']:
                        if r['source_row'] not in imported_rows:
                            self._record_import(conn, r, 'skip', None, now)
                            inserted['import_records'] += 1
                            imported_rows.add(r['source_row'])
                    continue

                # 行级已导入（理论上不会发生，因为桌级幂等已覆盖；防御）
                pending = [r for r in g['rows'] if r['source_row'] not in imported_rows]
                if not pending:
                    inserted['skipped_rows'] += len(g['rows'])
                    continue

                sid = g['source_id']
                # 机器分配：同(日期,机型)循环
                key = (str(g['date']), g['machine_type'])
                idx = day_counter[key]
                day_counter[key] += 1
                machine_id = (EIGHT_PORT_IDS if g['machine_type'] == '8port'
                              else FOUR_PORT_IDS)[idx % 2]

                # 时间：仅精确到日期
                start_time = f"{g['date']}T00:00:00"
                end_time = f"{g['date']}T23:59:59"
                total_fee = round(sum(r['amount'] or 0 for r in pending), 2)

                heads = [r for r in pending if r['is_table_head']]
                head_info = ''
                if heads:
                    h = heads[0]
                    head_info = f"桌首:{h['nickname']}; 桌首组织者:{h['table_head_organizer'] or '无'}; "
                orgs = list(set(r['organizer'] for r in pending if r['organizer']))
                if orgs:
                    head_info += f"组织者:{','.join(orgs)}"
                game_types = list(set(r['game_type'] for r in pending if r['game_type']))
                if game_types:
                    head_info += f"局型:{','.join(game_types)}; "

                note = (f"从Excel原始组局记录导入(legacy_excel_aug2026)。{head_info}"
                        f"time_precision=date_only").strip()
                import_quality = 'PARTIAL' if g['is_partial'] else 'OK'

                cur = conn.execute(
                    """INSERT INTO sessions (
                        machine_id, start_time, end_time, duration_minutes,
                        fee, final_fee, payment_method, status, note, source_id,
                        time_precision, import_quality
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    [machine_id, start_time, end_time, None,
                     total_fee, total_fee, 'legacy_unknown', 'completed', note, sid,
                     'date_only', import_quality]
                )
                session_id = cur.lastrowid
                inserted['sessions'] += 1

                for r in pending:
                    is_org = 1 if (r['player_id'] and r['nickname'] == r['organizer']) else 0
                    sp_cur = conn.execute(
                        """INSERT INTO session_players (
                            session_id, player_name, is_organizer, visit_type,
                            player_id, start_time, end_time, duration_minutes,
                            fee, final_fee, discount_amount, product_total,
                            grand_total, payment_method, status, is_overnight
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        [session_id, r['nickname'], is_org, 'active',
                         r['player_id'], start_time, end_time, None,
                         0.0, r['amount'] or 0.0, 0.0, 0.0, r['amount'] or 0.0,
                         'legacy_unknown', 'completed', r['is_overnight']]
                    )
                    sp_id = sp_cur.lastrowid
                    inserted['session_players'] += 1

                    # visit_records：按 (date, table_no, nickname) 去重
                    vr = conn.execute(
                        """SELECT id FROM visit_records WHERE visit_date=? AND table_number=?
                           AND player_name=?""",
                        [str(g['date']), g['table_no'], r['nickname']]
                    ).fetchone()
                    if vr:
                        conn.execute(
                            'UPDATE visit_records SET session_id=? WHERE id=?',
                            [session_id, vr['id']]
                        )
                    else:
                        conn.execute(
                            """INSERT INTO visit_records (
                                player_id, player_name, visit_date, machine_type,
                                game_type, brought_guest, organizer_name, is_overnight,
                                table_number, is_table_head, table_head_organizer,
                                data_quality, created_at, payment_amount, session_id
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                            [r['player_id'], r['nickname'], str(g['date']),
                             r['machine_type'], r['game_type'], r['brought_guest'],
                             r['organizer'], r['is_overnight'], g['table_no'],
                             r['is_table_head'], r['table_head_organizer'],
                             r['data_quality'], now, r['amount'] or 0.0, session_id]
                        )
                        inserted['visit_records'] += 1

                    # payments：仅有效金额
                    if r['amount'] is not None:
                        otn = (f"LEG-{str(g['date']).replace('-', '')}-{g['table_no']}-"
                               f"{r['player_id'] or 'U'}-{row_hash(r)[:8]}")
                        conn.execute(
                            """INSERT INTO payments (
                                out_trade_no, method, amount, status, provider,
                                session_player_id, created_at, updated_at
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                            [otn, 'legacy_unknown', r['amount'], 'SUCCESS',
                             SOURCE, sp_id, now, now]
                        )
                        inserted['payments'] += 1

                    # 行级登记
                    self._record_import(conn, r, 'session_player', sp_id, now)
                    inserted['import_records'] += 1
                    imported_rows.add(r['source_row'])

            # ---- 校验 ----
            checks = self._verify(conn, stats)
            ok = all(checks.values())
            if not ok:
                conn.execute('ROLLBACK')
                print('IMPORT_ROLLED_BACK')
                for k, v in checks.items():
                    if not v:
                        print(f'  [fail] {k}')
                return stats, None, False, checks
            conn.execute('COMMIT')
            print('IMPORT_COMMITTED')
            return stats, inserted, True, checks
        finally:
            conn.close()

    def _record_import(self, conn, r, entity_type, entity_id, now):
        h = row_hash({k: v for k, v in r.items() if not k.startswith('_')})
        try:
            conn.execute(
                """INSERT INTO legacy_import_records (
                    source, source_sheet, source_row, legacy_table_no, date,
                    entity_type, entity_id, imported_at, hash
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                [SOURCE, SHEET_NAME, r['source_row'], str(r['table_no']),
                 str(r['date']), entity_type, entity_id, now, h]
            )
        except sqlite3.IntegrityError:
            pass  # 已登记，幂等

    def _verify(self, conn, stats):
        """执行后校验（事务内）。返回 dict[name]=bool"""
        aug_prefix = 'excel_2026-08-'
        res = {}

        # 本次应插入 N 个新 session（would_insert_sessions）
        new_sessions = conn.execute(
            """SELECT COUNT(*) c FROM sessions
               WHERE source_id IN ({})""".format(
                ','.join('?' * len(stats['would_insert_groups']))
            ) if stats['would_insert_groups'] else 'SELECT 0 c',
            [g['source_id'] for g in stats['would_insert_groups']]
        ).fetchone()['c']
        res['session_count'] = new_sessions == stats['would_insert_sessions']

        # 8月历史 sessions 总数（含既有导入）
        aug_sessions = conn.execute(
            "SELECT COUNT(*) c FROM sessions WHERE source_id LIKE ?", [aug_prefix + '%']
        ).fetchone()['c']
        res['aug_session_total'] = aug_sessions == (stats['existing_aug_sessions']
                                                    + stats['would_insert_sessions'])

        # GMV：本次新增 payments 合计 == Excel 对应金额（严格，DIFF=0）
        gmv_new = conn.execute(
            """SELECT COALESCE(SUM(p.amount), 0) c FROM payments p
               JOIN session_players sp ON p.session_player_id = sp.id
               JOIN sessions s ON sp.session_id = s.id
               WHERE s.source_id IN ({})""".format(
                ','.join('?' * len(stats['would_insert_groups']))
            ) if stats['would_insert_groups'] else 'SELECT 0 c',
            [g['source_id'] for g in stats['would_insert_groups']]
        ).fetchone()['c']
        res['gmv_new'] = round(gmv_new, 2) == round(stats['would_insert_gmv'], 2)

        # 8月总 GMV（含 8/1-8/13 旧导入）== Excel GMV
        # 旧导入若有缺口（legacy_gap），默认 ROLLBACK；--allow-legacy-gap 显式放行
        total_gmv = conn.execute(
            """SELECT COALESCE(SUM(p.amount), 0) c FROM payments p
               JOIN session_players sp ON p.session_player_id = sp.id
               JOIN sessions s ON sp.session_id = s.id
               WHERE s.source_id LIKE ? AND p.amount IS NOT NULL""",
            [aug_prefix + '%']
        ).fetchone()['c']
        legacy_gap = round(stats['excel_gmv'] - round(total_gmv, 2), 2)
        res['gmv_total'] = (legacy_gap == 0) or self.allow_legacy_gap
        if not res['gmv_total']:
            print(f'  [legacy_gap] 8/1-8/13 旧导入与 Excel 缺口 {legacy_gap} 元，'
                  f'如需放行请加 --allow-legacy-gap')

        # 本次新增 payments 数
        new_payments = conn.execute(
            """SELECT COUNT(*) c FROM payments p
               JOIN session_players sp ON p.session_player_id = sp.id
               JOIN sessions s ON sp.session_id = s.id
               WHERE s.source_id IN ({})""".format(
                ','.join('?' * len(stats['would_insert_groups']))
            ) if stats['would_insert_groups'] else 'SELECT 0 c',
            [g['source_id'] for g in stats['would_insert_groups']]
        ).fetchone()['c']
        res['payment_count'] = new_payments == stats['would_insert_payments']

        # 本次新增 session_players 数
        new_sps = conn.execute(
            """SELECT COUNT(*) c FROM session_players
               WHERE session_id IN (SELECT id FROM sessions WHERE source_id IN ({}))""".format(
                ','.join('?' * len(stats['would_insert_groups']))
            ) if stats['would_insert_groups'] else 'SELECT 0 c',
            [g['source_id'] for g in stats['would_insert_groups']]
        ).fetchone()['c']
        res['session_player_count'] = new_sps == stats['would_insert_session_players']

        # 不修改 8/21 以后数据：导入前后 8/21+ sessions 数应一致（导入前为基线）
        after_821 = conn.execute(
            """SELECT COUNT(*) c FROM sessions
               WHERE start_time >= '2026-08-21' AND source_id IS NULL"""
        ).fetchone()['c']
        res['no_post_aug21_touch'] = after_821 == stats.get('_baseline_post_821', after_821)

        return res


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv=None):
    ap = argparse.ArgumentParser(description='QCOS 2026年8月历史组局/收银数据迁移导入器')
    ap.add_argument('--excel', required=True, help='源 Excel 路径')
    ap.add_argument('--db', default=os.path.join(ROOT, 'qcos.db'), help='目标 SQLite 库路径')
    ap.add_argument('--dry-run', action='store_true', help='只分析并生成报告，不写库')
    ap.add_argument('--execute', action='store_true', help='正式写入（需先备份成功）')
    ap.add_argument('--report-dir', default=os.path.join(ROOT, 'reports', 'aug2026_import'),
                    help='报告输出目录')
    ap.add_argument('--create-missing-players', action='store_true',
                    help='按 exact 标准昵称创建缺失玩家（默认不创建）')
    ap.add_argument('--allow-legacy-gap', action='store_true',
                    help='放行 8/1-8/13 旧导入与 Excel 的 GMV 缺口（默认缺口≠0 则回滚）')
    ap.add_argument('--ambiguous-bind', default='',
                    help='人工确认的重名绑定：昵称:player_id 逗号分隔，'
                         '如 back:84,BACK:84,坤哥:85,宁缺:110（大小写不敏感）')
    args = ap.parse_args(argv)

    if not os.path.exists(args.excel):
        print(f'Excel 不存在: {args.excel}')
        return 2
    if args.dry_run and args.execute:
        print('--dry-run 与 --execute 互斥')
        return 2

    # 解析 --ambiguous-bind
    ambiguous_bind = {}
    if args.ambiguous_bind:
        for item in args.ambiguous_bind.split(','):
            if ':' in item:
                k, v = item.split(':', 1)
                ambiguous_bind[k.strip()] = int(v.strip())

    imp = Importer(args.excel, args.db, args.report_dir,
                   create_missing=args.create_missing_players,
                   allow_legacy_gap=args.allow_legacy_gap,
                   ambiguous_bind=ambiguous_bind)
    try:
        stats = imp.analyze()
        summary = imp.write_reports(stats)
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        print(f'\n[报告] 已写入 {args.report_dir}/')

        if stats['bad_rows']:
            print(f'\n[bad_rows] {len(stats["bad_rows"])} 行：')
            for b in stats['bad_rows'][:15]:
                print(f"  行{b['source_row']} {b['date']} {b['nickname'] or '(空)'}: "
                      f"{'; '.join(b['issues'])}")
        if stats['unmatched_players']:
            print(f'\n[UNMATCHED_PLAYERS] {len(stats["unmatched_players"])} 人：')
            for nick, rows in sorted(stats['unmatched_players'].items()):
                print(f"  {nick}: 行{rows}")
        if stats['ambiguous_players']:
            print(f'\n[AMBIGUOUS_PLAYERS] {len(stats["ambiguous_players"])} 人（不自动绑定）：')
            for nick, rows in sorted(stats['ambiguous_players'].items()):
                print(f"  {nick}: 行{rows}")
        if stats['missing_payments']:
            print(f'\n[PAYMENT_MISSING] {len(stats["missing_payments"])} 行无有效金额')

        if args.dry_run:
            print('\nIMPORTER_DRY_RUN_COMPLETE')
            return 0

        if args.execute:
            print('\n[execute] 开始正式导入…')
            stats2, inserted, ok, checks = imp.execute()
            if not ok:
                print('IMPORT_FAILED')
                return 1
            print(f'  新增 sessions={inserted["sessions"]} '
                  f'session_players={inserted["session_players"]} '
                  f'payments={inserted["payments"]} '
                  f'visit_records={inserted["visit_records"]}')
            print('IMPORTER_EXECUTE_SUCCESS')
            return 0
        return 0
    finally:
        imp.loader.close()
        imp.db.close()


if __name__ == '__main__':
    sys.exit(main())