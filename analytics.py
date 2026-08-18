"""青鹭QCOS V1.5 数据分析中心 + AI经营报告导出系统
(Data Analytics Center & AI Report Exporter)

职责（来自任务）：
    数据聚合（营业 / 桌局 / 客户 / 运营四类指标）、时间筛选、
    指标计算、报告生成、数据导出（JSON 分析包 / Excel 经营分析）。

设计原则：
    1. 纯只读聚合 —— 不写库、不修改任何已有收银 / 计费 / 会员 / CRM /
       运营大脑 / 智能组局 / 桌局反馈 业务逻辑。
    2. 独立模块 —— 与 V1.1 operations、V1.2 player_matching、V1.3
       table_learning / operation_feedback 完全兼容（仅 consumer，无循环依赖）。
    3. 复用既有查询函数与评分结果（players 表中已沉淀的 customer_level /
       initiative_level / experience_score 等），不重复实现评分模型。

对外函数：
    resolve_range(filter_type, custom_start, custom_end) -> (start, end)        时间筛选
    get_dashboard(db, filter_type, custom_start, custom_end)  -> dict           实时指标
    build_business_summary(db)                                -> dict           business_summary.json
    build_players_analysis(db)                                -> dict           players_analysis.json
    build_table_analysis(db)                                  -> dict           table_analysis.json
    build_customer_segments(db)                               -> dict           customer_segments.json
    build_operation_tasks(db)                                 -> dict           operation_tasks.json
    build_monthly_report(db)                                  -> dict           monthly_report.json
    export_json_package(db, date_str)                        -> (bytes, name)  qcos_ai_report_*.zip
    export_excel(db, filter_type, custom_start, custom_end)  -> (bytes, name)  青鹭经营分析.xlsx
    AI_ANALYSIS_PROMPT (常量)                                                  AI_ANALYSIS_PROMPT.md 内容
"""

import io
import zipfile
import json
from datetime import datetime, date, timedelta
from calendar import monthrange

import operations
import player_matching as pm
import table_learning as tl
from operation_feedback import get_feedback


# ===================== 工具 =====================

def _now():
    return datetime.now()


def _today():
    return date.today()


def _parse_dt(s):
    if not s:
        return None
    s = str(s).strip()
    if 'T' in s:
        try:
            return datetime.fromisoformat(s)
        except ValueError:
            pass
    try:
        return datetime.strptime(s, '%Y-%m-%d')
    except ValueError:
        return None


def _gmv_in_range(db, start, end):
    """区间内 GMV = 台费实收(session_players.grand_total) + 商品销售(product_sales)。
    口径与 operations._gmv_in_range / /api/daily 保持一致（不含会员充值）。"""
    sids = [s['id'] for s in db.execute(
        "SELECT id FROM sessions WHERE status='closed' AND date(start_time) >= ? AND date(start_time) <= ?",
        [start, end]
    ).fetchall()]
    fee_total = 0.0
    if sids:
        sp = db.execute(
            'SELECT SUM(grand_total) t, SUM(final_fee) f FROM session_players WHERE session_id IN (%s)'
            % ','.join('?' * len(sids)), sids
        ).fetchone()
        fee_total = float(sp['t'] or sp['f'] or 0)
    prod = db.execute(
        'SELECT SUM(total) t FROM product_sales WHERE date(created_at) >= ? AND date(created_at) <= ?',
        [start, end]
    ).fetchone()['t'] or 0
    return round(float(fee_total) + float(prod), 2)


def _get_target(db):
    r = db.execute("SELECT value FROM settings WHERE key='monthly_target_gmv'").fetchone()
    if r:
        try:
            return float(r['value'])
        except (ValueError, TypeError):
            return 30000.0
    return 30000.0


def _historical_payment(db, start, end):
    """区间内 历史组局实收 = visit_records.payment_amount 之和（来自 Excel 导入的真实支付金额）。"""
    r = db.execute(
        "SELECT SUM(payment_amount) t FROM visit_records "
        "WHERE date(visit_date) >= ? AND date(visit_date) <= ? AND payment_amount > 0",
        [start, end]
    ).fetchone()
    return round(float(r['t'] or 0), 2)


# ===================== 时间筛选 =====================

def resolve_range(filter_type, custom_start=None, custom_end=None):
    """返回 (start_str, end_str)，均为 date.isoformat() 字符串（含边界）。"""
    today = _today()
    if filter_type == 'today':
        s = e = today.isoformat()
    elif filter_type == '7d':
        s = (today - timedelta(days=6)).isoformat()
        e = today.isoformat()
    elif filter_type == '30d':
        s = (today - timedelta(days=29)).isoformat()
        e = today.isoformat()
    elif filter_type == 'month':
        s = today.replace(day=1).isoformat()
        e = today.isoformat()
    elif filter_type == 'custom':
        s = (custom_start or today.isoformat())
        e = (custom_end or today.isoformat())
    else:
        s = e = today.isoformat()
    return s, e


# ===================== 四类指标聚合 =====================

def aggregate_business(db):
    """营业指标（固定参考：今日 / 本周 / 本月 / 目标 / 预测）。不受时间筛选影响。"""
    today = _today()
    t_str = today.isoformat()
    week_start = (today - timedelta(days=today.weekday())).isoformat()
    month_start = today.replace(day=1).isoformat()
    target = _get_target(db)

    today_gmv = _gmv_in_range(db, t_str, t_str)
    week_gmv = _gmv_in_range(db, week_start, t_str)
    month_gmv = _gmv_in_range(db, month_start, t_str)

    completion_pct = round(month_gmv / target * 100, 1) if target else 0.0
    remaining = max(0, target - month_gmv)

    d30 = (today - timedelta(days=30)).isoformat()
    last30 = _gmv_in_range(db, d30, t_str)
    avg_daily = round(last30 / 30, 2) if last30 else 0.0
    remain_days = (today.replace(day=monthrange(today.year, today.month)[1]) - today).days + 1
    forecast_month_end = round(month_gmv + avg_daily * remain_days, 2)
    forecast_gap = round(target - forecast_month_end, 2)

    historical_payment_month = _historical_payment(db, month_start, t_str)
    historical_payment_today = _historical_payment(db, t_str, t_str)

    return {
        'today_gmv': today_gmv,
        'week_gmv': week_gmv,
        'month_gmv': month_gmv,
        'month_target': target,
        'month_completion_pct': completion_pct,
        'month_remaining': round(remaining, 2),
        'avg_daily_gmv': avg_daily,
        'forecast_month_end': forecast_month_end,
        'forecast_gap': forecast_gap,
        'remain_days': remain_days,
        'historical_payment_today': historical_payment_today,
        'historical_payment_month': historical_payment_month,
    }


def aggregate_tables(db, start, end):
    """桌局指标（受时间筛选影响）。"""
    sessions = db.execute(
        "SELECT * FROM sessions WHERE date(start_time) >= ? AND date(start_time) <= ?",
        [start, end]
    ).fetchall()
    total = len(sessions)
    closed = [s for s in sessions if s['status'] == 'closed']

    durations = [s['duration_minutes'] for s in closed if s['duration_minutes']]
    avg_duration = round(sum(durations) / len(durations), 1) if durations else 0.0

    port_8 = sum(1 for s in sessions if s['machine_id'] in (1, 2))
    port_4 = sum(1 for s in sessions if s['machine_id'] in (3, 4))
    overnight_sids = set(r['session_id'] for r in db.execute(
        "SELECT DISTINCT session_id FROM session_players WHERE is_overnight=1"
    ).fetchall())
    overnight = sum(1 for s in sessions if s['id'] in overnight_sids)

    gmv = _gmv_in_range(db, start, end)
    avg_ticket = round(gmv / total, 2) if total else 0.0

    port_8_ratio = round(port_8 / total * 100, 1) if total else 0.0
    port_4_ratio = round(port_4 / total * 100, 1) if total else 0.0
    overnight_ratio = round(overnight / total * 100, 1) if total else 0.0

    return {
        'total_sessions': total,
        'closed_sessions': len(closed),
        'avg_duration_min': avg_duration,
        'avg_ticket': avg_ticket,
        'gmv': gmv,
        'port_8_count': port_8,
        'port_4_count': port_4,
        'port_8_ratio': port_8_ratio,
        'port_4_ratio': port_4_ratio,
        'overnight_count': overnight,
        'overnight_ratio': overnight_ratio,
    }


def _player_spend(db, pid, pname):
    """聚合单个玩家的累计消费（台费+商品）。"""
    pname = pname or ''
    sp_fee = db.execute(
        '''SELECT SUM(sp.grand_total) total FROM session_players sp
           JOIN sessions s ON sp.session_id = s.id
           WHERE (sp.player_id = ? OR sp.player_name = ?)''',
        [pid, pname]
    ).fetchone()['total'] or 0
    prod = db.execute(
        '''SELECT SUM(ps.total) total FROM product_sales ps
           JOIN session_players sp ON ps.session_player_id = sp.id
           JOIN sessions s ON sp.session_id = s.id
           WHERE (sp.player_id = ? OR sp.player_name = ?)''',
        [pid, pname]
    ).fetchone()['total'] or 0
    return round(float(sp_fee) + float(prod), 2)


def aggregate_customers(db, start, end):
    """客户指标。total/active_30d/level_a/level_b/churned 为全局参考；
    new_players / repeat_players 受时间筛选影响（基于 [start, end]）。"""
    players = db.execute('SELECT * FROM players').fetchall()
    total = len(players)

    d30 = (_today() - timedelta(days=30)).isoformat()
    active_30d = 0
    for p in players:
        p = dict(p)
        last = p.get('last_visit')
        if last and str(last) >= d30:
            active_30d += 1
        elif not last:
            cnt = db.execute(
                '''SELECT COUNT(*) c FROM session_players sp JOIN sessions s ON sp.session_id=s.id
                   WHERE (sp.player_id=? OR sp.player_name=?) AND date(s.start_time) >= ?''',
                [p['id'], p.get('name') or '', d30]
            ).fetchone()['c'] or 0
            if cnt > 0:
                active_30d += 1

    new_players = db.execute(
        '''SELECT p.id FROM players p
           WHERE (SELECT MIN(date(s.start_time)) FROM session_players sp
                  JOIN sessions s ON sp.session_id=s.id WHERE sp.player_id=p.id) BETWEEN ? AND ?''',
        [start, end]
    ).fetchall()
    new_count = len(new_players)

    repeat_players = db.execute(
        '''SELECT p.id FROM players p
           WHERE (SELECT COUNT(DISTINCT sp.session_id) FROM session_players sp
                  JOIN sessions s ON sp.session_id=s.id
                  WHERE (sp.player_id=p.id OR sp.player_name=p.name)
                    AND date(s.start_time) <= ?) >= 2''',
        [end]
    ).fetchall()
    repeat_count = len(repeat_players)

    churned = db.execute(
        '''SELECT id, name, customer_level, last_visit FROM players
           WHERE customer_level = 'D'
              OR (last_visit IS NOT NULL AND last_visit < ?)''',
        [d30]
    ).fetchall()
    churned_count = len(churned)

    level_a = sum(1 for p in players if dict(p).get('customer_level') in ('A+', 'A'))
    level_b = sum(1 for p in players if dict(p).get('customer_level') == 'B')

    return {
        'total_players': total,
        'active_30d': active_30d,
        'new_players': new_count,
        'repeat_players': repeat_count,
        'churned_players': churned_count,
        'level_a_count': level_a,
        'level_b_count': level_b,
    }


def aggregate_operations(db):
    """运营指标（全局）。任务完成率基于全部任务；组合数量实时计算。"""
    total_tasks = db.execute("SELECT COUNT(*) c FROM operation_tasks").fetchone()['c']
    done_tasks = db.execute(
        "SELECT COUNT(*) c FROM operation_tasks WHERE status='completed'").fetchone()['c']
    completion_rate = round(done_tasks / total_tasks * 100, 1) if total_tasks else 0.0

    players = db.execute('SELECT * FROM players').fetchall()
    active_init = sum(1 for p in players if dict(p).get('initiative_level') == 'active')
    passive_init = sum(1 for p in players if dict(p).get('initiative_level') == 'passive')

    organizer_candidates = db.execute(
        """SELECT COUNT(*) c FROM players
           WHERE is_organizer = 1
              OR (organizer_candidate IS NOT NULL AND organizer_candidate != ''
                  AND organizer_candidate NOT IN ('no', '否', '0'))"""
    ).fetchone()['c']

    best = tl.get_best_combinations(db, 999)
    risk = tl.get_risk_combinations(db, 999)

    return {
        'task_completion_rate': completion_rate,
        'task_total': total_tasks,
        'task_done': done_tasks,
        'active_initiative': active_init,
        'passive_initiative': passive_init,
        'organizer_candidates': organizer_candidates,
        'best_combinations_count': len(best),
        'risk_combinations_count': len(risk),
    }


def get_dashboard(db, filter_type='today', custom_start=None, custom_end=None):
    """实时经营概览聚合（供 /api/analytics/dashboard 与页面使用）。"""
    start, end = resolve_range(filter_type, custom_start, custom_end)
    return {
        'filter': {
            'type': filter_type,
            'start': start,
            'end': end,
            'label': {
                'today': '今日', '7d': '近7天', '30d': '近30天',
                'month': '本月', 'custom': '自定义'
            }.get(filter_type, '今日'),
        },
        'business': aggregate_business(db),
        'historical_payment_range': _historical_payment(db, start, end),
        'tables': aggregate_tables(db, start, end),
        'customers': aggregate_customers(db, start, end),
        'operations': aggregate_operations(db),
        'generated_at': _now().isoformat(timespec='seconds'),
    }


# ===================== AI 分析 JSON 构建 =====================

def build_business_summary(db):
    """business_summary.json：GMV / 桌数 / 客流 / 增长趋势。"""
    today = _today()
    t_str = today.isoformat()
    month_start = today.replace(day=1).isoformat()
    d30 = (today - timedelta(days=30)).isoformat()

    daily = []
    cur = _parse_dt(d30)
    while cur.date() <= today:
        ds = cur.isoformat()
        daily.append({'date': ds, 'gmv': _gmv_in_range(db, ds, ds)})
        cur += timedelta(days=1)

    daily_sessions = []
    cur = _parse_dt(d30)
    while cur.date() <= today:
        ds = cur.isoformat()
        cnt = db.execute(
            "SELECT COUNT(*) c FROM sessions WHERE date(start_time)=?", [ds]
        ).fetchone()['c']
        daily_sessions.append({'date': ds, 'sessions': cnt})
        cur += timedelta(days=1)

    biz = aggregate_business(db)
    return {
        'generated_at': _now().isoformat(timespec='seconds'),
        'today_gmv': biz['today_gmv'],
        'week_gmv': biz['week_gmv'],
        'month_gmv': biz['month_gmv'],
        'month_target': biz['month_target'],
        'month_completion_pct': biz['month_completion_pct'],
        'avg_daily_gmv': biz['avg_daily_gmv'],
        'forecast_month_end': biz['forecast_month_end'],
        'forecast_gap': biz['forecast_gap'],
        'total_sessions_month': db.execute(
            "SELECT COUNT(*) c FROM sessions WHERE date(start_time) >= ? AND date(start_time) <= ?",
            [month_start, t_str]).fetchone()['c'],
        'total_sessions_30d': db.execute(
            "SELECT COUNT(*) c FROM sessions WHERE date(start_time) >= ? AND date(start_time) <= ?",
            [d30, t_str]).fetchone()['c'],
        'daily_gmv_trend_30d': daily,
        'daily_sessions_30d': daily_sessions,
        'growth_note': (
            '若 forecast_month_end >= month_target 则增长趋势良好；'
            '若 forecast_gap > 0 则需提升日均GMV或加大引流。'
        ),
    }


def build_players_analysis(db):
    """players_analysis.json：玩家 等级 / 活跃 / 消费 / 组织能力 / 风险 / 体验评分。"""
    players = db.execute('SELECT * FROM players ORDER BY id').fetchall()
    rows = []
    for p in players:
        p = dict(p)
        pid = p['id']
        pname = p.get('name') or ''
        d30 = (_today() - timedelta(days=30)).isoformat()
        visits_30 = db.execute(
            '''SELECT COUNT(DISTINCT sp.session_id) c FROM session_players sp
               JOIN sessions s ON sp.session_id=s.id
               WHERE (sp.player_id=? OR sp.player_name=?) AND date(s.start_time) >= ?''',
            [pid, pname, d30]
        ).fetchone()['c'] or 0
        total_visits = db.execute(
            '''SELECT COUNT(DISTINCT sp.session_id) c FROM session_players sp
               JOIN sessions s ON sp.session_id=s.id
               WHERE (sp.player_id=? OR sp.player_name=?)''',
            [pid, pname]
        ).fetchone()['c'] or 0
        brought = db.execute(
            "SELECT COUNT(*) c FROM visit_records WHERE (player_id=? OR player_name=?) AND brought_guest=1",
            [pid, pname]
        ).fetchone()['c'] or 0
        spend = _player_spend(db, pid, pname)
        rows.append({
            'id': pid,
            'name': pname,
            'customer_level': p.get('customer_level') or 'C',
            'customer_score': p.get('customer_score'),
            'initiative_level': p.get('initiative_level') or 'unknown',
            'table_style': p.get('table_style_preference') or 'unknown',
            'visits_30d': visits_30,
            'total_visits': total_visits,
            'spend_total': spend,
            'brought_guest': int(brought),
            'is_organizer': p.get('is_organizer') or 0,
            'organizer_candidate': p.get('organizer_candidate') or '',
            'experience_score': p.get('experience_score'),
            'compatibility_score': p.get('compatibility_score'),
            'conflict_count': p.get('conflict_count') or 0,
            'risk_tags': p.get('risk_tags') or '',
            'maintenance_priority': p.get('maintenance_priority') or '',
            'last_visit': p.get('last_visit'),
        })
    return {
        'generated_at': _now().isoformat(timespec='seconds'),
        'player_count': len(rows),
        'players': rows,
    }


def build_table_analysis(db):
    """table_analysis.json：历史桌局 / 组合 / 评分 / 反馈 / 优秀组合 / 风险组合。"""
    sessions = db.execute(
        "SELECT * FROM sessions ORDER BY start_time DESC LIMIT 200"
    ).fetchall()
    session_list = []
    for s in sessions:
        s = dict(s)
        players = db.execute(
            'SELECT DISTINCT p.name FROM session_players sp '
            'LEFT JOIN players p ON p.id=sp.player_id WHERE sp.session_id=? AND sp.player_id IS NOT NULL',
            [s['id']]
        ).fetchall()
        fb = db.execute('SELECT id FROM session_feedback WHERE session_id=?', [s['id']]).fetchone()
        session_list.append({
            'session_id': s['id'],
            'start_time': s['start_time'],
            'status': s['status'],
            'duration_minutes': s['duration_minutes'],
            'fee': s.get('fee'),
            'final_fee': s.get('final_fee'),
            'players': [p['name'] for p in players],
            'has_feedback': bool(fb),
        })

    pairs = db.execute("SELECT * FROM player_pair_stats ORDER BY play_count DESC").fetchall()
    pair_list = []
    for r in pairs:
        r = dict(r)
        a = db.execute('SELECT name FROM players WHERE id=?', [r['player_a_id']]).fetchone()
        b = db.execute('SELECT name FROM players WHERE id=?', [r['player_b_id']]).fetchone()
        pair_list.append({
            'player_a': a['name'] if a else r['player_a_id'],
            'player_b': b['name'] if b else r['player_b_id'],
            'play_count': r['play_count'],
            'positive_count': r['positive_count'],
            'negative_count': r['negative_count'],
            'average_score': r['average_score'],
            'relationship_trend': r['relationship_trend'],
            'last_play_date': r['last_play_date'],
        })

    fbs = get_feedback(db, limit=100)
    feedback_list = []
    for f in fbs:
        f = dict(f)
        feedback_list.append({
            'session_id': f['session_id'],
            'start_time': f.get('start_time'),
            'atmosphere_score': f['atmosphere_score'],
            'compatibility_score': f['compatibility_score'],
            'table_quality_score': f['table_quality_score'],
            'average_score': round((f['atmosphere_score'] + f['compatibility_score'] + f['table_quality_score']) / 3.0, 2),
            'conflict_level': f['conflict_level'],
            'conflict_type': f['conflict_type'],
            'notes': f.get('notes'),
            'player_names': f.get('player_names'),
        })

    best = tl.get_best_combinations(db, 50)
    risk = tl.get_risk_combinations(db, 50)

    return {
        'generated_at': _now().isoformat(timespec='seconds'),
        'sessions': session_list,
        'player_pairs': pair_list,
        'feedbacks': feedback_list,
        'best_combinations': best,
        'risk_combinations': risk,
    }


def build_customer_segments(db):
    """customer_segments.json：A/B/C/D 客户分层。"""
    players = db.execute('SELECT * FROM players ORDER BY id').fetchall()
    segments = {'A+': [], 'A': [], 'B': [], 'C': [], 'D': []}
    counts = {}
    for p in players:
        p = dict(p)
        lvl = p.get('customer_level') or 'C'
        if lvl not in segments:
            lvl = 'C'
        counts[lvl] = counts.get(lvl, 0) + 1
        spend = _player_spend(db, p['id'], p.get('name') or '')
        segments[lvl].append({
            'id': p['id'],
            'name': p.get('name'),
            'customer_score': p.get('customer_score'),
            'spend_total': spend,
            'last_visit': p.get('last_visit'),
            'risk_tags': p.get('risk_tags') or '',
            'maintenance_priority': p.get('maintenance_priority') or '',
        })
    return {
        'generated_at': _now().isoformat(timespec='seconds'),
        'level_labels': operations.LEVEL_LABELS,
        'counts': {
            'A+': counts.get('A+', 0), 'A': counts.get('A', 0),
            'B': counts.get('B', 0), 'C': counts.get('C', 0),
            'D': counts.get('D', 0),
        },
        'segments': segments,
    }


def build_operation_tasks(db):
    """operation_tasks.json：当前未完成任务（按类型分组）。"""
    rows = db.execute(
        '''SELECT t.*, p.name as player_name, p.customer_level
           FROM operation_tasks t LEFT JOIN players p ON t.player_id=p.id
           WHERE t.status='pending'
           ORDER BY CASE t.priority WHEN 'high' THEN 0 WHEN 'medium' THEN 1 ELSE 2 END, t.created_at'''
    ).fetchall()
    out = {}
    for r in rows:
        r = dict(r)
        out.setdefault(r['task_type'], []).append({
            'id': r['id'],
            'player_id': r['player_id'],
            'player_name': r.get('player_name'),
            'customer_level': r.get('customer_level'),
            'priority': r['priority'],
            'description': r['description'],
            'created_at': r['created_at'],
        })
    return {
        'generated_at': _now().isoformat(timespec='seconds'),
        'task_type_labels': operations.TASK_TYPE_LABELS,
        'pending_total': len(rows),
        'tasks_by_type': out,
    }


def build_monthly_report(db):
    """monthly_report.json：月度经营数据。"""
    today = _today()
    t_str = today.isoformat()
    month_start = today.replace(day=1).isoformat()
    biz = aggregate_business(db)
    tbl = aggregate_tables(db, month_start, t_str)
    cust = aggregate_customers(db, month_start, t_str)
    ops = aggregate_operations(db)
    best = tl.get_best_combinations(db, 20)
    risk = tl.get_risk_combinations(db, 20)

    machine_revenue = []
    for mid, mname in [(1, '八口机1'), (2, '八口机2'), (3, '四口机1'), (4, '四口机2')]:
        sids = [s['id'] for s in db.execute(
            "SELECT id FROM sessions WHERE machine_id=? AND status='closed' AND date(start_time)>=? AND date(start_time)<=?",
            [mid, month_start, t_str]
        ).fetchall()]
        fee = 0.0
        if sids:
            sp = db.execute(
                'SELECT SUM(grand_total) t FROM session_players WHERE session_id IN (%s)'
                % ','.join('?' * len(sids)), sids
            ).fetchone()
            fee = float(sp['t'] or 0)
        machine_revenue.append({'machine': mname, 'gmv': round(fee, 2), 'sessions': len(sids)})

    return {
        'generated_at': _now().isoformat(timespec='seconds'),
        'month': month_start[:7],
        'gmv': biz['month_gmv'],
        'target': biz['month_target'],
        'completion_pct': biz['month_completion_pct'],
        'forecast_month_end': biz['forecast_month_end'],
        'forecast_gap': biz['forecast_gap'],
        'sessions': tbl['total_sessions'],
        'avg_duration_min': tbl['avg_duration_min'],
        'avg_ticket': tbl['avg_ticket'],
        'port_8_ratio': tbl['port_8_ratio'],
        'port_4_ratio': tbl['port_4_ratio'],
        'overnight_ratio': tbl['overnight_ratio'],
        'machine_revenue': machine_revenue,
        'total_players': cust['total_players'],
        'active_30d': cust['active_30d'],
        'new_players': cust['new_players'],
        'repeat_players': cust['repeat_players'],
        'churned_players': cust['churned_players'],
        'level_a_count': cust['level_a_count'],
        'level_b_count': cust['level_b_count'],
        'task_completion_rate': ops['task_completion_rate'],
        'best_combinations': best,
        'risk_combinations': risk,
    }


# ===================== 导出：AI 分析包 (ZIP + JSON) =====================

def export_json_package(db, date_str=None):
    """生成 qcos_ai_report_YYYYMMDD.zip，内含 6 个 JSON + AI分析提示文件。
    返回 (bytes, filename)。"""
    if not date_str:
        date_str = _today().strftime('%Y%m%d')

    files = {
        'business_summary.json': build_business_summary(db),
        'players_analysis.json': build_players_analysis(db),
        'table_analysis.json': build_table_analysis(db),
        'customer_segments.json': build_customer_segments(db),
        'operation_tasks.json': build_operation_tasks(db),
        'monthly_report.json': build_monthly_report(db),
        'AI_ANALYSIS_PROMPT.md': AI_ANALYSIS_PROMPT,
    }

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fname, content in files.items():
            if fname.endswith('.json'):
                data = json.dumps(content, ensure_ascii=False, indent=2)
            else:
                data = content
            zf.writestr(fname, data)
    buf.seek(0)
    return buf.getvalue(), f'qcos_ai_report_{date_str}.zip'


# ===================== 导出：Excel 经营分析 =====================

def export_excel(db, filter_type='month', custom_start=None, custom_end=None):
    """生成 青鹭经营分析.xlsx（9 个 sheet）。返回 (bytes, filename)。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    start, end = resolve_range(filter_type, custom_start, custom_end)
    wb = Workbook()

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2C5F2D')
    title_font = Font(bold=True, size=13)

    def style_header(ws, ncols, row=1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    def write_table(ws, headers, rows, start_row=1):
        for j, h in enumerate(headers, 1):
            ws.cell(row=start_row, column=j, value=h)
        style_header(ws, len(headers), start_row)
        for i, row in enumerate(rows, start_row + 1):
            for j, val in enumerate(row, 1):
                ws.cell(row=i, column=j, value=val)
        for j, h in enumerate(headers, 1):
            col = chr(64 + j) if j <= 26 else 'A'
            ws.column_dimensions[col].width = max(12, min(40, len(str(h)) + 4))

    # ---- Sheet 1: 经营总览 ----
    ws = wb.active
    ws.title = '经营总览'
    ws.cell(row=1, column=1, value='青鹭雀庄 经营分析总览').font = title_font
    ws.cell(row=2, column=1, value=f'生成时间：{_now().isoformat(timespec="seconds")}')
    ws.cell(row=3, column=1, value=f'统计区间：{start} ~ {end}')
    biz = aggregate_business(db)
    tbl = aggregate_tables(db, start, end)
    cust = aggregate_customers(db, start, end)
    ops = aggregate_operations(db)
    overview = [
        ('今日GMV', biz['today_gmv']),
        ('本周GMV', biz['week_gmv']),
        ('本月GMV', biz['month_gmv']),
        ('月目标', biz['month_target']),
        ('月目标完成率(%)', biz['month_completion_pct']),
        ('平均每日GMV', biz['avg_daily_gmv']),
        ('预计月底GMV', biz['forecast_month_end']),
        ('预计缺口', biz['forecast_gap']),
        ('区间总桌数', tbl['total_sessions']),
        ('区间平均桌时长(分)', tbl['avg_duration_min']),
        ('区间平均客单价', tbl['avg_ticket']),
        ('八口占比(%)', tbl['port_8_ratio']),
        ('四口占比(%)', tbl['port_4_ratio']),
        ('通宵占比(%)', tbl['overnight_ratio']),
        ('玩家总数', cust['total_players']),
        ('30天活跃玩家', cust['active_30d']),
        ('新增玩家', cust['new_players']),
        ('复购玩家', cust['repeat_players']),
        ('流失玩家', cust['churned_players']),
        ('A级玩家', cust['level_a_count']),
        ('B级玩家', cust['level_b_count']),
        ('任务完成率(%)', ops['task_completion_rate']),
        ('主动型玩家', ops['active_initiative']),
        ('被动型玩家', ops['passive_initiative']),
        ('常务候选', ops['organizer_candidates']),
        ('优秀组合数', ops['best_combinations_count']),
        ('风险组合数', ops['risk_combinations_count']),
    ]
    write_table(ws, ['指标', '数值'], overview, start_row=5)

    # ---- Sheet 2: 每日GMV ----
    ws = wb.create_sheet('每日GMV')
    daily = []
    cur = _parse_dt(start)
    end_d = _parse_dt(end)
    if cur and end_d:
        while cur <= end_d:
            ds = cur.isoformat()
            cnt = db.execute("SELECT COUNT(*) c FROM sessions WHERE date(start_time)=?", [ds]).fetchone()['c']
            pl = db.execute(
                '''SELECT COUNT(DISTINCT sp.player_id) c FROM session_players sp
                   JOIN sessions s ON sp.session_id=s.id WHERE date(s.start_time)=?''', [ds]
            ).fetchone()['c']
            daily.append([ds, _gmv_in_range(db, ds, ds), cnt, pl])
            cur += timedelta(days=1)
    write_table(ws, ['日期', 'GMV', '桌数', '玩家数'], daily)

    # ---- Sheet 3: 玩家列表 ----
    ws = wb.create_sheet('玩家列表')
    players = db.execute('SELECT * FROM players ORDER BY id').fetchall()
    prows = []
    for p in players:
        p = dict(p)
        pid = p['id']
        pname = p.get('name') or ''
        spend = _player_spend(db, pid, pname)
        prows.append([
            pid, pname, p.get('customer_level') or 'C', p.get('customer_score'),
            p.get('initiative_level') or 'unknown', p.get('table_style_preference') or 'unknown',
            p.get('visits_30d'), p.get('total_visits'), spend,
            p.get('is_organizer') or 0, p.get('experience_score'),
            p.get('compatibility_score'), p.get('conflict_count') or 0,
            p.get('risk_tags') or '', p.get('last_visit'),
        ])
    write_table(ws, ['ID', '姓名', '等级', '价值分', '主动性', '局型偏好', '30天到店',
                     '累计到店', '累计消费', '组织者', '体验分', '适配分', '冲突数', '风险标签', '最近到店'],
                prows)

    # ---- Sheet 4: 客户分层 ----
    ws = wb.create_sheet('客户分层')
    seg = build_customer_segments(db)
    srows = []
    for lvl in ['A+', 'A', 'B', 'C', 'D']:
        for m in seg['segments'][lvl]:
            srows.append([lvl, seg['level_labels'].get(lvl, lvl), m.get('name'),
                          m.get('customer_score'), m.get('spend_total'),
                          m.get('last_visit'), m.get('risk_tags') or '', m.get('maintenance_priority') or ''])
    write_table(ws, ['等级', '等级标签', '姓名', '价值分', '累计消费', '最近到店', '风险标签', '维护优先级'], srows)

    # ---- Sheet 5: 桌局记录 ----
    ws = wb.create_sheet('桌局记录')
    sessions = db.execute("SELECT * FROM sessions ORDER BY start_time DESC LIMIT 500").fetchall()
    trows = []
    for s in sessions:
        s = dict(s)
        players_names = db.execute(
            'SELECT DISTINCT p.name FROM session_players sp LEFT JOIN players p ON p.id=sp.player_id '
            'WHERE sp.session_id=? AND sp.player_id IS NOT NULL', [s['id']]
        ).fetchall()
        m = db.execute('SELECT name FROM machines WHERE id=?', [s['machine_id']]).fetchone()
        trows.append([
            s['id'], s['start_time'], s['status'], m['name'] if m else s['machine_id'],
            s.get('duration_minutes'), s.get('fee'), s.get('final_fee'),
            ', '.join(p['name'] for p in players_names),
        ])
    write_table(ws, ['桌局ID', '开始时间', '状态', '台桌', '时长(分)', '台费', '实收', '玩家'], trows)

    # ---- Sheet 6: 玩家关系 ----
    ws = wb.create_sheet('玩家关系')
    rels = db.execute('SELECT * FROM player_relationships ORDER BY id').fetchall()
    rrows = []
    for r in rels:
        r = dict(r)
        a = db.execute('SELECT name FROM players WHERE id=?', [r['player_a_id']]).fetchone()
        b = db.execute('SELECT name FROM players WHERE id=?', [r['player_b_id']]).fetchone()
        rrows.append([
            r['id'], a['name'] if a else r['player_a_id'], b['name'] if b else r['player_b_id'],
            r['relationship_type'], r['relationship_score'], r['source'],
            r.get('note') or '', r.get('updated_at'),
        ])
    write_table(ws, ['ID', '玩家A', '玩家B', '关系', '评分', '来源', '备注', '更新时间'], rrows)

    # ---- Sheet 7: 优秀组合 ----
    ws = wb.create_sheet('优秀组合')
    best = tl.get_best_combinations(db, 100)
    brows = []
    for c in best:
        brows.append([
            ' + '.join(c['player_names']), c['play_count'], c['average_score'],
            c['recommend'],
        ])
    write_table(ws, ['玩家组合', '同组次数', '平均评分', '推荐度'], brows)

    # ---- Sheet 8: 风险组合 ----
    ws = wb.create_sheet('风险组合')
    risk = tl.get_risk_combinations(db, 100)
    krows = []
    for c in risk:
        krows.append([
            c['player_a_name'], c['player_b_name'], c['play_count'],
            c['average_score'], c['negative_count'], c['positive_count'],
            c['reason'], c['suggestion'],
        ])
    write_table(ws, ['玩家A', '玩家B', '同组次数', '平均评分', '负向次数', '正向次数', '原因', '建议'], krows)

    # ---- Sheet 9: 任务完成 ----
    ws = wb.create_sheet('任务完成')
    tasks = db.execute(
        '''SELECT t.*, p.name as player_name FROM operation_tasks t
           LEFT JOIN players p ON t.player_id=p.id ORDER BY t.created_at DESC LIMIT 500'''
    ).fetchall()
    krows = []
    for t in tasks:
        t = dict(t)
        krows.append([
            t['id'], operations.TASK_TYPE_LABELS.get(t['task_type'], t['task_type']),
            t.get('player_name'), t['priority'], t['status'],
            t.get('description'), t['created_at'], t.get('completed_at'),
        ])
    write_table(ws, ['ID', '类型', '玩家', '优先级', '状态', '描述', '创建时间', '完成时间'], krows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), f'青鹭经营分析_{_today().strftime("%Y%m%d")}.xlsx'


# ===================== AI 分析提示文件 =====================

AI_ANALYSIS_PROMPT = """# 青鹭雀庄 AI 经营分析提示词

你是「青鹭雀庄」（成都九眼桥阳光新业日麻k庄）的经营分析顾问。
本目录由 QCOS 系统导出的 **AI分析包** 构成，用于对雀庄经营进行客观、可执行的分析。

## 数据文件说明

| 文件 | 内容 |
|------|------|
| `business_summary.json` | 今日/本周/本月 GMV、月目标完成率、近30天每日GMV趋势、桌数、客流 |
| `players_analysis.json` | 每位玩家的等级、30天/累计到店、累计消费、带客、组织者、风险、体验评分 |
| `table_analysis.json` | 历史桌局、玩家组合共现统计、桌局反馈、优秀组合、风险组合 |
| `customer_segments.json` | A+/A/B/C/D 客户分层人数与各层代表玩家 |
| `operation_tasks.json` | 当前未完成的运营任务（按类型分组） |
| `monthly_report.json` | 本月经营汇总（GMV/目标/完成率/各机型营收/客户分层/组合） |

> 评分口径：customer_level 由运营大脑计算（A+≥85 / A≥70 / B≥50 / C≥30 / D<30）；
> experience_score / compatibility_score 来自桌局反馈回流（0-100）；
> initiative_level 来自主动性模型（active/semi_active/passive/unknown）。

## 分析目标（请逐项回答）

1. **经营健康度判断**
   - 月目标完成率是否达标？预计月底GMV与缺口如何？
   - 客流结构是否健康（活跃/新增/复购/流失比例）？
   - 桌局质量（优秀/风险组合）是否稳定？

2. **GMV 增长空间**
   - 从近30天每日GMV趋势看，增长瓶颈在哪？
   - 哪个时段/机型（八口/四口）营收贡献高、可加大供给？
   - 沉默/流失客户（D级、30天未到）的召回可释放多少潜在GMV？

3. **核心客户识别**
   - 谁是「青鹭最有价值的核心生态玩家」？（高等级 + 高带客 + 组织者 + 正向体验）
   - 他们的消费与到店规律是什么？如何针对性维护？

4. **流失风险发现**
   - 哪些玩家处于流失边缘（D级 / 近30天未到 / 风险标签）？
   - 哪些组合存在历史冲突（风险组合），需避免同桌？

5. **运营建议**
   - 给出 3-5 条具体、可落地的下周运营动作（组局、召回、引流、定价、常务培养）。
   - 结合 operation_tasks.json 中未完成任务，指出优先级。

## 输出要求
- 用中文，分点清晰，给出量化依据（引用具体数值/玩家名）。
- 避免空泛套话；每条建议必须对应数据中的证据。
- 若某数据为空（如暂无桌局反馈），明确说明「数据不足，建议先补全反馈」。
"""
