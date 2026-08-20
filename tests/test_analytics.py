"""QCOS V1.5 数据分析中心 + AI 经营报告导出 - 单元测试

覆盖：
    1. 数据聚合（get_dashboard 四类指标 / GMV 计算口径）
    2. 日期筛选（resolve_range 五种类型边界）
    3. JSON 导出（AI 分析包 ZIP 含 7 文件、可解压、JSON 合法）
    4. Excel 导出（9 个 sheet / 可被 openpyxl 读取）
    5. 无数据情况（空库不崩溃、返回零值与空结构）
    6. 真实数据库读取（若存在 qcos.db，端到端跑通）

使用临时数据库（基于真实 schema），不污染生产库。
"""

import os
import sys
import io
import json
import zipfile
import tempfile
from datetime import datetime, date, timedelta

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, ROOT)

# ===== 测试专用环境变量（必须在 import config 之前设置，非生产凭据）=====
os.environ.setdefault('QCOS_SECRET_KEY', 'test_secret_key_qcos_unit_test')
os.environ.setdefault('QCOS_ADMIN_PASSWORD', 'test_admin_pw_qcos_v2')

import config
# 指向临时库，避免污染生产数据
_tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.db')
config.DB_PATH = _tmp.name

import models
models.init_db()
import analytics as A

DB = models.get_db()
DB.execute('PRAGMA busy_timeout=30000')

TODAY = date.today()
TODAY_STR = TODAY.isoformat()
REAL_DB = os.path.join(ROOT, 'qcos.db')

PASS = 0
FAIL = 0


def ok(name):
    global PASS
    PASS += 1
    print(f'  [PASS] {name}')


def bad(name, msg):
    global FAIL
    FAIL += 1
    print(f'  [FAIL] {name}: {msg}')


def check(name, cond, msg=''):
    if cond:
        ok(name)
    else:
        bad(name, msg)


# ===================== 测试夹具 =====================

def add_player(name, **kw):
    cols = ['name'] + list(kw.keys())
    vals = [name] + list(kw.values())
    ph = ','.join('?' * len(cols))
    cur = DB.execute(f"INSERT INTO players ({','.join(cols)}) VALUES ({ph})", vals)
    DB.commit()
    return cur.lastrowid


def add_session(pid_list, start=TODAY_STR + ' 20:00:00', machine_id=1, status='closed',
                duration=120, grand_total=0.0, final_fee=0.0, overnight=False):
    cur = DB.execute(
        'INSERT INTO sessions (machine_id, start_time, status, duration_minutes, fee, final_fee) '
        'VALUES (?, ?, ?, ?, ?, ?)',
        [machine_id, start, status, duration, final_fee, final_fee])
    sid = cur.lastrowid
    DB.commit()
    n = max(1, len(pid_list))
    per = round(grand_total / n, 2)
    per_f = round(final_fee / n, 2)
    for pid in pid_list:
        DB.execute(
            'INSERT INTO session_players (session_id, player_name, player_id, status, '
            'grand_total, final_fee, is_overnight) VALUES (?, ?, ?, "playing", ?, ?, ?)',
            [sid, 'p', pid, per, per_f, 1 if overnight else 0])
    DB.commit()
    return sid


def add_product(sid, spid, total=0.0, created_at=None):
    if created_at is None:
        created_at = TODAY_STR + ' 21:00:00'
    DB.execute(
        'INSERT INTO product_sales (session_id, session_player_id, product_name, price, quantity, total, created_at) '
        'VALUES (?, ?, ?, ?, 1, ?, ?)',
        [sid, spid, '测试商品', total, total, created_at])
    DB.commit()


def add_task(player_id, task_type='recover_customer', status='pending'):
    DB.execute(
        'INSERT INTO operation_tasks (player_id, task_type, priority, status, created_at) '
        'VALUES (?, ?, "normal", ?, ?)',
        [player_id, task_type, status, TODAY_STR])
    DB.commit()


def seed_demo_data():
    """插入一套可验证的演示数据。"""
    # 玩家：A级组织者 / B级 / D级流失 / 普通
    p1 = add_player('阿一', customer_level='A', initiative_level='active', is_organizer=1,
                    customer_score=88, last_visit=TODAY_STR, total_visits=20, visits_30d=8,
                    experience_score=90, compatibility_score=85)
    p2 = add_player('阿二', customer_level='B', initiative_level='semi_active',
                    last_visit=TODAY_STR, total_visits=10, visits_30d=3)
    p3 = add_player('阿三', customer_level='D', initiative_level='passive',
                    last_visit=(TODAY - timedelta(days=60)).isoformat(), total_visits=2, visits_30d=0)
    p4 = add_player('阿四', customer_level='C', initiative_level='unknown',
                    last_visit=(TODAY - timedelta(days=5)).isoformat(), total_visits=5, visits_30d=1)

    # 今日一桌八口机：台费200 + 商品50
    sid = add_session([p1, p2, p3, p4], machine_id=1, grand_total=200.0, final_fee=200.0)
    add_product(sid, p1, total=50.0)

    # 昨日四口机一桌（区间：近7天/近30天/本月）
    sid2 = add_session([p1, p4], start=(TODAY - timedelta(days=1)).isoformat() + ' 21:00:00',
                       machine_id=3, grand_total=40.0, final_fee=40.0)

    # 通宵桌
    sid3 = add_session([p2, p3], start=(TODAY - timedelta(days=2)).isoformat() + ' 23:30:00',
                       machine_id=2, grand_total=100.0, final_fee=100.0, overnight=True)

    # 运营任务：1 完成 / 2 未完成
    add_task(p1, 'recover_customer', 'completed')
    add_task(p2, 'maintain_customer', 'pending')
    add_task(p3, 'risk_warning', 'pending')

    return [p1, p2, p3, p4]


# ===================== 1. 数据聚合 =====================

def test_aggregation():
    print('\n[1] 数据聚合')
    seed_demo_data()
    d = A.get_dashboard(DB, 'today')
    check('dashboard 含四类指标',
          all(k in d for k in ('business', 'tables', 'customers', 'operations', 'filter')),
          f'keys={list(d.keys())}')
    check('filter 类型正确', d['filter']['type'] == 'today', d['filter'])
    check('今日GMV=台费+商品',
          abs(d['business']['today_gmv'] - 250.0) < 0.01,
          f"today_gmv={d['business']['today_gmv']}")
    check('月目标完成率>0', d['business']['month_completion_pct'] > 0,
          f"pct={d['business']['month_completion_pct']}")
    check('桌局-总桌数(今日)=1', d['tables']['total_sessions'] == 1,
          f"total={d['tables']['total_sessions']}")
    check('桌局-八口占比=100%', d['tables']['port_8_ratio'] == 100.0,
          f"ratio={d['tables']['port_8_ratio']}")
    check('客户-玩家总数=4', d['customers']['total_players'] == 4,
          f"total={d['customers']['total_players']}")
    check('客户-A级数量=1', d['customers']['level_a_count'] == 1,
          f"a={d['customers']['level_a_count']}")
    check('客户-流失预警=1(仅D级)', d['customers']['churned_players'] == 1,
          f"churned={d['customers']['churned_players']}")
    check('运营-任务完成率=33.3%',
          abs(d['operations']['task_completion_rate'] - 33.3) < 0.2,
          f"rate={d['operations']['task_completion_rate']}")
    check('运营-主动型=1', d['operations']['active_initiative'] == 1,
          f"active={d['operations']['active_initiative']}")
    check('运营-常务候选=1', d['operations']['organizer_candidates'] == 1,
          f"org={d['operations']['organizer_candidates']}")


# ===================== 2. 日期筛选 =====================

def test_date_filter():
    print('\n[2] 日期筛选')
    base = date.today()
    s, e = A.resolve_range('today')
    check('today 边界', (s, e) == (base.isoformat(), base.isoformat()), f'{s},{e}')
    s, e = A.resolve_range('7d')
    check('7d 起点=today-6', s == (base - timedelta(days=6)).isoformat(), s)
    check('7d 终点=today', e == base.isoformat(), e)
    s, e = A.resolve_range('30d')
    check('30d 起点=today-29', s == (base - timedelta(days=29)).isoformat(), s)
    s, e = A.resolve_range('month')
    check('month 起点=月初', s == base.replace(day=1).isoformat(), s)
    check('month 终点=today', e == base.isoformat(), e)
    s, e = A.resolve_range('custom', '2026-01-01', '2026-01-31')
    check('custom 透传', (s, e) == ('2026-01-01', '2026-01-31'), f'{s},{e}')

    # 30d 筛选下应聚合到全部演示桌（共3桌：今日/昨日/2天前）
    d = A.get_dashboard(DB, '30d')
    check('30d 总桌数=3', d['tables']['total_sessions'] == 3,
          f"total={d['tables']['total_sessions']}")
    check('30d 通宵占比=(1/3)', abs(d['tables']['overnight_ratio'] - 33.3) < 0.2,
          f"overnight={d['tables']['overnight_ratio']}")


# ===================== 3. JSON 导出 =====================

def test_json_export():
    print('\n[3] JSON 导出 (AI 分析包)')
    blob, name = A.export_json_package(DB, '20260814')
    check('返回 bytes', isinstance(blob, bytes) and len(blob) > 0, f'len={len(blob)}')
    check('文件名格式', name == 'qcos_ai_report_20260814.zip', name)
    check('ZIP 魔数 PK', blob[:2] == b'PK', blob[:4])
    try:
        zf = zipfile.ZipFile(io.BytesIO(blob))
        names = zf.namelist()
        expected = ['business_summary.json', 'players_analysis.json', 'table_analysis.json',
                    'customer_segments.json', 'operation_tasks.json', 'monthly_report.json',
                    'AI_ANALYSIS_PROMPT.md']
        missing = [n for n in expected if n not in names]
        check('ZIP 含 7 个文件', not missing, f'missing={missing}; got={names}')
        for n in expected:
            if n.endswith('.json'):
                raw = zf.read(n).decode('utf-8')
                json.loads(raw)  # 必须合法 JSON
        check('各 JSON 均合法', True)
        md = zf.read('AI_ANALYSIS_PROMPT.md').decode('utf-8')
        check('提示词含分析目标', '分析目标' in md, '缺少分析目标')
    except Exception as ex:
        bad('ZIP 解析', str(ex))


# ===================== 4. Excel 导出 =====================

def test_excel_export():
    print('\n[4] Excel 导出')
    blob, name = A.export_excel(DB, 'month')
    check('返回 bytes', isinstance(blob, bytes) and len(blob) > 0, f'len={len(blob)}')
    check('文件名格式', name.startswith('青鹭经营分析') and name.endswith('.xlsx'), name)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(blob))
        sheets = wb.sheetnames
        expected = ['经营总览', '每日GMV', '玩家列表', '客户分层', '桌局记录',
                    '玩家关系', '优秀组合', '风险组合', '任务完成']
        check('9 个 sheet 齐全', sheets == expected, f'got={sheets}')
        # 玩家列表应含 4 名玩家
        ws = wb['玩家列表']
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        rows = [r for r in rows if r and r[0] is not None]
        check('玩家列表行数=4', len(rows) == 4, f'rows={len(rows)}')
        # 桌局记录应含 3 桌
        ws2 = wb['桌局记录']
        rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
        rows2 = [r for r in rows2 if r and r[0] is not None]
        check('桌局记录行数=3', len(rows2) == 3, f'rows={len(rows2)}')
    except Exception as ex:
        bad('Excel 解析', str(ex))


# ===================== 5. 无数据情况 =====================

def test_empty_db():
    print('\n[5] 无数据情况')
    # 全新空库
    tmp2 = tempfile.NamedTemporaryFile(delete=False, suffix='.db')
    old_cfg = config.DB_PATH
    old_mod = models.DB_PATH
    config.DB_PATH = tmp2.name
    models.DB_PATH = tmp2.name
    models.init_db()
    db2 = models.get_db()
    db2.execute('PRAGMA busy_timeout=30000')
    try:
        d = A.get_dashboard(db2, '30d')
        check('空库 dashboard 不崩溃', d is not None and 'business' in d)
        check('空库 今日GMV=0', d['business']['today_gmv'] == 0, d['business']['today_gmv'])
        check('空库 玩家总数=0', d['customers']['total_players'] == 0)
        check('空库 风险组合=0', d['operations']['risk_combinations_count'] == 0)
        blob, name = A.export_json_package(db2, '20260101')
        zf = zipfile.ZipFile(io.BytesIO(blob))
        check('空库 ZIP 仍含 7 文件', len(zf.namelist()) == 7, zf.namelist())
        blob2, _ = A.export_excel(db2, 'month')
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(blob2))
        check('空库 Excel 仍 9 sheet', len(wb.sheetnames) == 9, wb.sheetnames)
    except Exception as ex:
        bad('空库导出', str(ex))
    finally:
        db2.close()
        config.DB_PATH = old_cfg
        models.DB_PATH = old_mod
        try:
            os.remove(tmp2.name)
        except OSError:
            pass


# ===================== 6. 真实数据库读取 =====================

def test_real_db():
    print('\n[6] 真实数据库读取')
    if not os.path.exists(REAL_DB):
        print('  [SKIP] 未找到真实库 qcos.db，跳过')
        return
    # 真实库只读副本 + 跑一次 init_db 迁移（含 players.status 归档字段），
    # 验证新 schema 下 analytics 对真实数据端到端可用；绝不直接改真实库。
    import shutil
    real_copy = tempfile.NamedTemporaryFile(delete=False, suffix='.db')
    real_copy.close()
    shutil.copy2(REAL_DB, real_copy.name)
    old_cfg = config.DB_PATH
    old_mod = models.DB_PATH
    config.DB_PATH = real_copy.name
    models.DB_PATH = real_copy.name
    try:
        models.init_db()
        rdb = models.get_db()
        rdb.execute('PRAGMA busy_timeout=30000')
        d = A.get_dashboard(rdb, '30d')
        check('真实库 dashboard 结构完整',
              all(k in d for k in ('business', 'tables', 'customers', 'operations')),
              f'keys={list(d.keys())}')
        blob, _ = A.export_json_package(rdb, '20260814')
        zf = zipfile.ZipFile(io.BytesIO(blob))
        check('真实库 ZIP 含 7 文件', len(zf.namelist()) == 7, zf.namelist())
        blob2, _ = A.export_excel(rdb, '30d')
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(blob2))
        check('真实库 Excel 9 sheet', len(wb.sheetnames) == 9, wb.sheetnames)
        rdb.close()
    except Exception as ex:
        bad('真实库读取', str(ex))
    finally:
        config.DB_PATH = old_cfg
        models.DB_PATH = old_mod
        try:
            os.remove(real_copy.name)
        except OSError:
            pass


if __name__ == '__main__':
    print('=' * 60)
    print('QCOS V1.5 analytics 测试')
    print('=' * 60)
    test_aggregation()
    test_date_filter()
    test_json_export()
    test_excel_export()
    test_empty_db()
    test_real_db()
    print('\n' + '=' * 60)
    print(f'结果：通过 {PASS} / 失败 {FAIL}')
    print('=' * 60)
    sys.exit(1 if FAIL else 0)
